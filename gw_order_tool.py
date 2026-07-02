import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import openpyxl
import calendar
import json
import math
import subprocess
import sys
import os
import threading
from datetime import date, timedelta


class GWOrderTool:
    UNMATCHED_LAST_WEEK_DESC = (
        "Unmatched Product Codes — ordered last week but missing from this week's pad "
        "(barcode or product code may have changed)"
    )
    UNMATCHED_NETO_DESC = (
        "Unmatched Product Codes — sold on Neto but not found in this week's table "
        "(barcode or product code may have changed; ignore or reassign to restock the right line)"
    )
    NETO_QTY_SOLD_HEADER = "Qty Sold on Neto"
    QTY_IN_PACK_HEADER = "Units per Pack"
    PACKS_TO_ORDER_HEADER = "Packs to Order"
    SELLABLE_STOCK_HEADER = "Sellable Stock"
    RECOMMENDED_PACKS_HEADER = "Recommended Packs to Order"
    # Sellable stock on Neto at or below this means the shelf is running low and one
    # pack should be reordered — see _recommended_packs().
    LOW_STOCK_THRESHOLD = 2
    PACK_REVIEW_DESC = (
        "Pack-Based Products Needing Review — sold in packs of more than one. "
        "Recommended Packs to Order is based on current sellable stock (1 pack when sellable "
        "stock is 2 or less), not on qty sold. Sellable Stock shows on-hand stock in "
        "brackets when they differ. Tick one or more lines, right-click to change "
        "their recommendation, then Confirm to send them back to the main table, or "
        "Ignore to skip ordering them."
    )
    SAVED_PRODUCTS_DESC = (
        "Carried Over From Last Run — saved from a previous unmatched list for later "
        "follow-up; select and Remove once handled"
    )
    SAVED_PRODUCTS_FILENAME = "saved_products.json"

    # Bulk-select checkbox column, prepended to every table by _populate_tree.
    # It's a live view of the tree's normal selection: clicking a row's box toggles
    # that row in/out of the selection without needing Ctrl/Shift, and clicking the
    # column heading ticks/unticks all visible rows. Every action button then simply
    # operates on the ticked (selected) rows. See _on_tree_click/_sync_check_marks.
    CHECK_HEADER = "☑"
    CHECKED = "☑"
    UNCHECKED = "☐"

    # Used by _resize_tree_columns() to proportionally fill a table's full width instead
    # of leaving blank space — ttk's built-in column "stretch" is unreliable across
    # platforms (e.g. doesn't kick in consistently on macOS), so widths are recomputed
    # by hand on every resize. Product Name gets the lion's share of any extra space;
    # the checkbox column stays at a fixed sliver.
    COLUMN_WEIGHTS = {"Product Name": 3, CHECK_HEADER: 0}
    DEFAULT_COLUMN_WEIGHT = 1
    COLUMN_MIN_WIDTHS = {"Product Name": 200, "Recommended Packs to Order": 160, CHECK_HEADER: 34}
    DEFAULT_COLUMN_MIN_WIDTH = 90

    # Initial pixel width of the left column (Unmatched/Pack Review) — kept narrow and
    # fixed (see _sync_paned_layout) so the main table dominates the window.
    LEFT_PANE_WIDTH = 420

    def __init__(self, root):
        self.root = root
        self.root.title("GW Order Tool")
        self.root.geometry("1200x700")
        self.root.minsize(900, 500)

        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.saved_products_path = os.path.join(self.script_dir, self.SAVED_PRODUCTS_FILENAME)
        self.saved_products = self._load_saved_products()
        self.saved_products_frame = None
        self.saved_products_tree = None
        self.saved_products_row_data = {}

        self.order_pad_path = tk.StringVar()
        self.order_conf_path = tk.StringVar()
        self.this_week_pad_path = tk.StringVar()
        self.order_pad_data = []
        self.conf_qty_lookup = {}
        self.outstanding_lookup = {}
        self.qty_in_pack_lookup = {}
        self.last_week_ordered_lookup = {}
        self.this_week_pad_data = []
        self.showing_outstanding_view = False
        self.paned = None
        self.left_paned = None
        self.unmatched_frame = None
        self.unmatched_tree = None
        self.unmatched_row_data = {}
        # Which main-table column an unmatched row's "qty" represents, and therefore
        # which column Reassign should write to — "Qty Outstanding" before Fetch Stock
        # from Neto has run, self.NETO_QTY_SOLD_HEADER after. Kept in sync by
        # _set_unmatched_rows().
        self.unmatched_target_column = "Qty Outstanding"
        self.main_table_frame = None
        self.pack_review_frame = None
        self.pack_review_tree = None
        self.pack_review_row_data = {}
        self.main_row_iid = {}
        # Sellable/on-hand stock per SKU as scraped from Neto — {sku: (available, on_hand)}.
        # Feeds the Sellable Stock column and the stock-based pack recommendation.
        self.neto_stock_lookup = {}
        # Original position of each product code in this week's pad, so a confirmed
        # Pack Review line can slot back into the main table where it came from.
        self.main_code_pad_index = {}
        # Carried Over From Last Run and Pack Review only become visible once Fetch
        # Stock from Neto has completed at least once — before that, only Unmatched
        # Product Codes is relevant (see _sync_paned_layout()).
        self.neto_fetch_done = False
        self.neto_from_date = self._get_last_tuesday()
        self.neto_btn = None

        self._setup_styles()
        self._build_top_bar()
        self._build_file_selection_frame()
        self.table_container = None
        self.tree = None
        self.bottom_frame = None
        self.status_label = None
        self.neto_progress = None
        self.status_var = tk.StringVar()

    @staticmethod
    def _get_last_tuesday():
        """Mirrors neto_scraper.get_last_tuesday() — kept in sync manually since the GUI
        launches that script as a separate subprocess rather than importing it.
        Tuesday of the week *before* the current one (weeks run Mon–Sun), e.g. run on
        Thursday 2/7 -> current week is Mon 29/6-Sun 5/7, so this returns 23/6."""
        today = date.today()
        monday_this_week = today - timedelta(days=today.weekday())
        return monday_this_week + timedelta(days=1) - timedelta(days=7)

    @staticmethod
    def _format_date(d):
        return d.strftime("%d/%m/%Y")

    def _load_saved_products(self):
        """Load products saved via 'Save for Next Order' in a previous session. Missing
        or corrupt file just means there's nothing carried over yet — not an error."""
        try:
            with open(self.saved_products_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                return data
        except (FileNotFoundError, json.JSONDecodeError, OSError):
            pass
        return []

    def _write_saved_products(self):
        try:
            with open(self.saved_products_path, "w", encoding="utf-8") as f:
                json.dump(self.saved_products, f, indent=2)
        except OSError as e:
            messagebox.showerror("Failed to Save", f"Couldn't write {self.SAVED_PRODUCTS_FILENAME}:\n{e}")

    def _setup_styles(self):
        style = ttk.Style()
        style.configure(
            "NetoChecking.TLabel",
            font=("TkDefaultFont", 10, "bold"),
            foreground="#b45309",
        )
        style.configure(
            "NetoLoginWait.TLabel",
            font=("TkDefaultFont", 10, "bold"),
            foreground="#b91c1c",
        )
        style.configure(
            "NetoDateSelected.TButton",
            font=("TkDefaultFont", 9, "bold"),
        )
        style.map(
            "NetoDateSelected.TButton",
            background=[("!disabled", "#2563eb")],
            foreground=[("!disabled", "#ffffff")],
        )
        # Minimalist collapse/expand arrow for the left-column panel headers — a plain
        # label rather than a full ttk.Button, so it doesn't look like a big chunky
        # button next to the panel description.
        style.configure(
            "PanelToggle.TLabel",
            font=("TkDefaultFont", 8),
            foreground="#9ca3af",
        )

    def _build_top_bar(self):
        bar = ttk.Frame(self.root)
        bar.pack(fill="x", padx=10, pady=(10, 0))
        ttk.Button(bar, text="Start Over", command=self._start_over).pack(side="right")

    def _build_file_selection_frame(self):
        frame = ttk.LabelFrame(self.root, text="Select Files", padding=10)
        frame.pack(fill="x", padx=10, pady=(8, 5))

        self.order_pad_label = ttk.Label(frame, text="Last Week's Order Pad (.xlsx):")
        self.order_pad_entry = ttk.Entry(frame, textvariable=self.order_pad_path, width=80)
        self.order_pad_browse_btn = ttk.Button(frame, text="Browse...", command=self._browse_order_pad)
        self.order_pad_label.grid(row=0, column=0, sticky="w", pady=2)
        self.order_pad_entry.grid(row=0, column=1, padx=5, pady=2)
        self.order_pad_browse_btn.grid(row=0, column=2, pady=2)

        self.order_conf_label = ttk.Label(frame, text="Last Week's Order Confirmation (.xlsx):")
        self.order_conf_entry = ttk.Entry(frame, textvariable=self.order_conf_path, width=80)
        self.order_conf_browse_btn = ttk.Button(frame, text="Browse...", command=self._browse_order_conf)
        self.order_conf_label.grid(row=1, column=0, sticky="w", pady=2)
        self.order_conf_entry.grid(row=1, column=1, padx=5, pady=2)
        self.order_conf_browse_btn.grid(row=1, column=2, pady=2)

        self.reconcile_btn = ttk.Button(frame, text="Reconcile Orders", command=self._on_compare)
        self.reconcile_btn.grid(row=2, column=1, pady=(10, 0))

        # Rows hidden once Compare Orders has successfully run — see _hide_last_week_rows().
        self.last_week_row_widgets = [
            self.order_pad_label, self.order_pad_entry, self.order_pad_browse_btn,
            self.order_conf_label, self.order_conf_entry, self.order_conf_browse_btn,
            self.reconcile_btn,
        ]

        self.this_week_label = ttk.Label(frame, text="This Week's Order Pad (.xlsx):")
        self.this_week_entry = ttk.Entry(frame, textvariable=self.this_week_pad_path, width=80)
        self.this_week_browse_btn = ttk.Button(frame, text="Browse...", command=self._browse_this_week_pad)
        self.compare_btn = ttk.Button(frame, text="Compare Orders", command=self._on_compare_this_week)
        # Hidden until Reconcile Orders has run; revealed by _reveal_this_week_row().
        self.this_week_row_widgets = [
            (self.this_week_label, dict(row=3, column=0, sticky="w", pady=2)),
            (self.this_week_entry, dict(row=3, column=1, padx=5, pady=2)),
            (self.this_week_browse_btn, dict(row=3, column=2, pady=2)),
            (self.compare_btn, dict(row=4, column=1, pady=(10, 0))),
        ]

        # Neto "Date Placed From" selector — sits to the left of the Fetch Stock from Neto
        # button. Built here but only gridded once the button is promoted (see
        # _promote_compare_to_fetch_button), since it's only relevant at that stage.
        self.neto_date_frame = ttk.Frame(frame)
        ttk.Label(self.neto_date_frame, text="Date Placed From:").pack(side="left", padx=(0, 4))
        self.neto_date_var = tk.StringVar(value=self._format_date(self.neto_from_date))
        self.neto_date_entry = ttk.Entry(
            self.neto_date_frame, textvariable=self.neto_date_var, width=11, state="readonly"
        )
        self.neto_date_entry.pack(side="left", padx=(0, 4))
        ttk.Button(self.neto_date_frame, text="Change...", command=self._open_date_picker).pack(side="left")

        frame.columnconfigure(1, weight=1)

    def _reveal_this_week_row(self):
        for widget, grid_opts in self.this_week_row_widgets:
            widget.grid(**grid_opts)

    def _hide_last_week_rows(self):
        for widget in self.last_week_row_widgets:
            widget.grid_remove()

    def _promote_compare_to_fetch_button(self):
        """Once Compare Orders has run, the same button slot becomes the Fetch Stock from Neto trigger."""
        self.compare_btn.config(text="Fetch Stock from Neto", command=self._on_fetch_neto, state="normal")
        self.neto_btn = self.compare_btn
        self.neto_date_frame.grid(row=4, column=0, sticky="e", padx=(0, 5), pady=(10, 0))

    def _start_over(self):
        """Reset everything back to the initial state — for when the wrong files got
        selected or the data otherwise needs to be redone from scratch."""
        if self.neto_btn is not None and self.neto_btn.cget("state") == "disabled":
            messagebox.showwarning(
                "Fetch in Progress",
                "A Neto stock check is still running. Please wait for it to finish before starting over.",
            )
            return

        if not messagebox.askyesno(
            "Start Over",
            "This clears all selected files and loaded data so you can start again. Continue?",
        ):
            return

        # Clear all loaded/derived data.
        self.order_pad_path.set("")
        self.order_conf_path.set("")
        self.this_week_pad_path.set("")
        self.order_pad_data = []
        self.conf_qty_lookup = {}
        self.outstanding_lookup = {}
        self.qty_in_pack_lookup = {}
        self.last_week_ordered_lookup = {}
        self.this_week_pad_data = []
        self.showing_outstanding_view = False
        self.paned = None
        self.left_paned = None
        self.unmatched_frame = None
        self.unmatched_tree = None
        self.unmatched_row_data = {}
        self.unmatched_target_column = "Qty Outstanding"
        self.main_table_frame = None
        self.pack_review_frame = None
        self.pack_review_tree = None
        self.pack_review_row_data = {}
        self.main_row_iid = {}
        self.neto_stock_lookup = {}
        self.main_code_pad_index = {}
        self.neto_fetch_done = False
        # Note: self.saved_products itself is NOT reset — it's persisted in
        # saved_products.json independently of this working session.
        self.saved_products_frame = None
        self.saved_products_tree = None
        self.saved_products_row_data = {}
        self.neto_from_date = self._get_last_tuesday()
        self.neto_date_var.set(self._format_date(self.neto_from_date))

        # Tear down the results area built up by Reconcile/Compare/Fetch.
        if self.table_container:
            self.table_container.destroy()
            self.table_container = None
        self.tree = None

        # Restore the file-selection stage to how it looked on first launch.
        for widget in self.last_week_row_widgets:
            widget.grid()
        self.reconcile_btn.config(state="normal")

        for widget, _opts in self.this_week_row_widgets:
            widget.grid_remove()
        self.compare_btn.config(text="Compare Orders", command=self._on_compare_this_week, state="normal")
        self.neto_date_frame.grid_remove()

        self._build_bottom_frame()
        self.status_var.set("Ready. Select last week's Order Pad and Order Confirmation to begin.")

    def _open_date_picker(self):
        """Small self-contained month-calendar popup for picking Neto's 'Date Placed From'
        filter — no extra dependency (like tkcalendar) required."""
        picker = tk.Toplevel(self.root)
        picker.title("Select Date Placed From")
        picker.transient(self.root)
        picker.grab_set()
        picker.resizable(False, False)

        view_year = tk.IntVar(value=self.neto_from_date.year)
        view_month = tk.IntVar(value=self.neto_from_date.month)

        header = ttk.Frame(picker)
        header.pack(fill="x", padx=8, pady=(8, 4))
        month_label = ttk.Label(header, text="", anchor="center", width=16)

        days_frame = ttk.Frame(picker)
        days_frame.pack(padx=8, pady=(0, 4))

        def on_pick(y, m, d):
            self.neto_from_date = date(y, m, d)
            self.neto_date_var.set(self._format_date(self.neto_from_date))
            picker.destroy()

        def refresh_calendar():
            for widget in days_frame.winfo_children():
                widget.destroy()
            y, m = view_year.get(), view_month.get()
            month_label.config(text=date(y, m, 1).strftime("%B %Y"))

            for col, h in enumerate(["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]):
                ttk.Label(days_frame, text=h, anchor="center", width=4).grid(row=0, column=col, padx=1, pady=1)

            cal = calendar.Calendar(firstweekday=0)
            for row, week in enumerate(cal.monthdayscalendar(y, m), start=1):
                for col, day in enumerate(week):
                    if day == 0:
                        ttk.Label(days_frame, text="", width=4).grid(row=row, column=col, padx=1, pady=1)
                        continue
                    is_selected = (
                        y == self.neto_from_date.year
                        and m == self.neto_from_date.month
                        and day == self.neto_from_date.day
                    )
                    # Selected day gets both a distinct style AND a text marker — ttk button
                    # background colors are unreliable on some platforms (e.g. macOS Aqua),
                    # so the bracket marker keeps the selection visible regardless of theme.
                    btn = ttk.Button(
                        days_frame,
                        text=f"[{day}]" if is_selected else str(day),
                        width=4,
                        style="NetoDateSelected.TButton" if is_selected else "TButton",
                        command=lambda d=day: on_pick(y, m, d),
                    )
                    btn.grid(row=row, column=col, padx=1, pady=1)

        def change_month(delta):
            y, m = view_year.get(), view_month.get() + delta
            if m < 1:
                y, m = y - 1, 12
            elif m > 12:
                y, m = y + 1, 1
            view_year.set(y)
            view_month.set(m)
            refresh_calendar()

        ttk.Button(header, text="◀", width=3, command=lambda: change_month(-1)).pack(side="left")
        month_label.pack(side="left", expand=True)
        ttk.Button(header, text="▶", width=3, command=lambda: change_month(1)).pack(side="left")

        footer = ttk.Frame(picker)
        footer.pack(fill="x", padx=8, pady=(0, 8))
        last_tue = self._get_last_tuesday()
        ttk.Button(
            footer, text="Last Tuesday",
            command=lambda: on_pick(last_tue.year, last_tue.month, last_tue.day),
        ).pack(side="left")
        ttk.Button(footer, text="Cancel", command=picker.destroy).pack(side="right")

        refresh_calendar()

    def _make_collapsible_panel(self, parent, description, on_toggle=None):
        """Wrap a panel in a manually collapsible header (▼/▶ toggle + description),
        independent of whether the panel currently has data. Used for the three
        left-column panels (Carried Over, Unmatched, Pack Review) so the user can show
        or hide any of them at will instead of relying on automatic show/hide.
        on_toggle, if given, is called after every toggle — used to reflow the sash
        positions so a collapsed panel actually frees up its space (see
        _update_left_pane_sizes) instead of leaving it blank."""
        outer = ttk.Frame(parent, padding=5)
        header = ttk.Frame(outer)
        header.pack(fill="x", pady=(0, 4))

        state = {"collapsed": False}

        # Plain, small label instead of a ttk.Button — a real button looks oversized
        # next to a one-line description; the whole header row is clickable anyway.
        toggle_btn = ttk.Label(header, text="▼", width=2, style="PanelToggle.TLabel", cursor="hand2")
        toggle_btn.pack(side="left")
        desc_label = ttk.Label(
            header, text=description, wraplength=380, justify="left", cursor="hand2"
        )
        desc_label.pack(side="left", fill="x", expand=True, padx=(4, 0))

        content = ttk.Frame(outer)
        content.pack(fill="both", expand=True)

        def toggle(event=None):
            state["collapsed"] = not state["collapsed"]
            if state["collapsed"]:
                content.pack_forget()
                toggle_btn.config(text="▶")
            else:
                content.pack(fill="both", expand=True)
                toggle_btn.config(text="▼")
            if on_toggle:
                on_toggle()

        def set_description(text):
            desc_label.config(text=text)

        outer.gw_is_collapsed = lambda: state["collapsed"]
        toggle_btn.bind("<Button-1>", toggle)
        desc_label.bind("<Button-1>", toggle)

        outer.gw_set_description = set_description
        return outer, content

    def _create_treeview(self, parent, expand=True, height=None, searchable=True):
        container = ttk.Frame(parent)
        container.pack(fill="both" if expand else "x", expand=expand)

        table_frame = ttk.Frame(container)
        tree = (
            ttk.Treeview(table_frame, show="headings", height=height)
            if height
            else ttk.Treeview(table_frame, show="headings")
        )

        # "extended" is ttk's default selectmode too, but set it explicitly so bulk
        # selection (Shift-click for a range, Ctrl/Cmd-click to toggle individual rows,
        # Ctrl/Cmd+A to select all) is guaranteed on every table regardless of platform
        # or theme defaults.
        tree.configure(selectmode="extended")
        tree.bind("<Control-a>", lambda e, t=tree: self._select_all_rows(t))
        tree.bind("<Command-a>", lambda e, t=tree: self._select_all_rows(t))

        # Checkbox-column interactions — box clicks toggle rows in/out of the
        # selection, heading click ticks/unticks everything visible, and any
        # selection change (including Ctrl/Shift-click or Ctrl+A) refreshes the
        # tick marks so they always mirror the actual selection.
        tree.bind("<Button-1>", lambda e, t=tree: self._on_tree_click(t, e))
        tree.bind("<<TreeviewSelect>>", lambda e, t=tree: self._sync_check_marks(t))

        if searchable:
            search_frame = ttk.Frame(container)
            search_frame.pack(fill="x", pady=(0, 4))
            ttk.Label(search_frame, text="Search:").pack(side="left", padx=(0, 4))
            search_var = tk.StringVar()
            ttk.Entry(search_frame, textvariable=search_var, width=30).pack(side="left")
            ttk.Button(search_frame, text="Clear", command=lambda: search_var.set("")).pack(
                side="left", padx=(4, 0)
            )
            # Stashed on the widget itself so _populate_tree/_filter_tree can find them
            # again without needing a separate registry keyed by tree.
            tree.gw_search_var = search_var
            search_var.trace_add("write", lambda *_args: self._filter_tree(tree))

        table_frame.pack(fill="both" if expand else "x", expand=expand)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        return tree

    def _browse_order_pad(self):
        path = filedialog.askopenfilename(
            title="Select Order Pad",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if path:
            self.order_pad_path.set(path)

    def _browse_order_conf(self):
        path = filedialog.askopenfilename(
            title="Select Order Confirmation",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if path:
            self.order_conf_path.set(path)

    def _browse_this_week_pad(self):
        path = filedialog.askopenfilename(
            title="Select This Week's Order Pad",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if path:
            self.this_week_pad_path.set(path)

    def _on_compare_this_week(self):
        path = self.this_week_pad_path.get().strip()
        if not path:
            messagebox.showwarning("Missing File", "Please select this week's Order Pad before comparing.")
            return

        try:
            self.this_week_pad_data = self._parse_order_pad(path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load This Week's Order Pad:\n{e}")
            return

        self._show_outstanding_view()
        self._hide_last_week_rows()
        self._promote_compare_to_fetch_button()

    def _on_compare(self):
        pad_path = self.order_pad_path.get().strip()
        conf_path = self.order_conf_path.get().strip()

        if not pad_path or not conf_path:
            messagebox.showwarning("Missing Files", "Please select both files before comparing.")
            return

        self.showing_outstanding_view = False
        self._load_order_pad(pad_path)
        self._load_order_conf(conf_path)
        self._refresh_view()

        self.reconcile_btn.config(state="disabled")
        self._reveal_this_week_row()

    def _parse_order_pad(self, path):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        data_rows = []
        for row in rows[4:]:
            product_code = row[5]
            if product_code is None or str(product_code).strip() == "":
                continue
            order_qty = row[6]
            try:
                qty = int(order_qty) if order_qty is not None and str(order_qty).strip() != "" else 0
            except (ValueError, TypeError):
                qty = 0
            product_name = row[7] if len(row) > 7 and row[7] is not None else ""

            # Column J — how many units come in one orderable pack (e.g. 6 sprays per
            # pack). Used to convert raw qty needed into whole packs to order.
            pack_raw = row[9] if len(row) > 9 else None
            try:
                qty_in_pack = int(pack_raw) if pack_raw is not None and str(pack_raw).strip() != "" else 1
            except (ValueError, TypeError):
                qty_in_pack = 1
            if qty_in_pack <= 0:
                qty_in_pack = 1

            data_rows.append((product_code, qty, product_name, qty_in_pack))

        return data_rows

    def _load_order_pad(self, path):
        try:
            self.order_pad_data = self._parse_order_pad(path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Order Pad:\n{e}")

    def _load_order_conf(self, path):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb["Table 1"]

            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            self.conf_qty_lookup = {}
            for row in rows[1:]:
                product_code = row[0]
                if product_code is None:
                    continue
                qty_available = row[4]
                try:
                    qty = int(qty_available) if qty_available is not None and str(qty_available).strip() != "" else 0
                except (ValueError, TypeError):
                    qty = 0
                self.conf_qty_lookup[str(product_code)] = qty

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Order Confirmation:\n{e}")

    def _refresh_view(self):
        if self.table_container:
            self.table_container.destroy()
        if self.bottom_frame:
            self.bottom_frame.destroy()

        self.table_container = ttk.Frame(self.root)
        self.table_container.pack(fill="both", expand=True, padx=10, pady=(5, 5))
        self.tree = self._create_treeview(self.table_container)
        self._enable_row_copy(self.tree)

        headers = ["Product Code", "Product Name", "Qty Ordered", "Qty Confirmed", "Qty Outstanding"]
        merged = []
        last_week_ordered_lookup = {}
        for product_code, qty_ordered, product_name, _qty_in_pack in self.order_pad_data:
            qty_confirmed = self.conf_qty_lookup.get(str(product_code), 0)
            # If more got confirmed than we ordered (e.g. confirmed without a matching
            # order), floor at 0 rather than showing a negative outstanding qty.
            qty_outstanding = max(qty_ordered - qty_confirmed, 0)
            merged.append((product_code, product_name, qty_ordered, qty_confirmed, qty_outstanding))
            if qty_ordered > 0:
                last_week_ordered_lookup[str(product_code)] = (product_code, product_name, qty_outstanding)
        self._populate_tree(self.tree, headers, merged)

        self.outstanding_lookup = {
            str(code): outstanding for code, _name, _qty_ordered, _qty_confirmed, outstanding in merged
        }
        self.last_week_ordered_lookup = last_week_ordered_lookup

        self.status_var.set("Reconciled. Select this week's Order Pad to continue.")
        self._build_bottom_frame()

    def _build_bottom_frame(self):
        if self.bottom_frame:
            self.bottom_frame.destroy()

        self.bottom_frame = ttk.Frame(self.root)
        self.bottom_frame.pack(fill="x", padx=10, pady=(0, 10))

        self.status_label = ttk.Label(self.bottom_frame, textvariable=self.status_var)
        self.status_label.pack(side="left")

        # Indeterminate progress bar shown only while actively checking against Neto —
        # see _set_neto_checking(). Not packed until needed.
        self.neto_progress = ttk.Progressbar(self.bottom_frame, mode="indeterminate", length=160)

    def _show_outstanding_view(self):
        if self.table_container:
            self.table_container.destroy()

        self.table_container = ttk.Frame(self.root)
        self.table_container.pack(fill="both", expand=True, padx=10, pady=(5, 5))

        this_week_codes = set()
        seen = set()
        matched_rows = []
        self.qty_in_pack_lookup = {}
        for product_code, _qty_ordered, product_name, qty_in_pack in self.this_week_pad_data:
            key = str(product_code)
            this_week_codes.add(key)
            if key in seen:
                continue
            seen.add(key)
            self.qty_in_pack_lookup[key] = qty_in_pack
            qty_outstanding = self.outstanding_lookup.get(key, 0)
            # Qty Sold on Neto and Packs to Order start blank — filled in once Fetch
            # Stock from Neto runs, see _apply_neto_stock_data(). Qty in Pack isn't
            # shown as its own column here — it's tracked in qty_in_pack_lookup and
            # only surfaces indirectly via Packs to Order (and directly in the Pack
            # Review panel for items that need manual checking).
            matched_rows.append((product_code, product_name, qty_outstanding, "", ""))

        # Codes that had a real order last week but don't appear in this week's pad at all —
        # likely because the barcode/product code or name changed.
        unmatched_rows = [
            (orig_code, product_name, qty_outstanding)
            for key, (orig_code, product_name, qty_outstanding) in self.last_week_ordered_lookup.items()
            if key not in this_week_codes
        ]

        # Outer horizontal split: left column stacks the Carried Over, Unmatched Product
        # Codes and Pack Review panels — all three are always attached (see
        # _sync_paned_layout()) and individually collapsible, rather than being
        # automatically shown/hidden based on whether they currently have data; the
        # main outstanding table sits on the right and is always present.
        self.paned = ttk.PanedWindow(self.table_container, orient="horizontal")
        self.paned.pack(fill="both", expand=True)

        self.left_paned = ttk.PanedWindow(self.paned, orient="vertical")

        # Carried Over From Last Run — products saved from a previous session's Unmatched
        # Product Codes list, loaded from saved_products.json. Sits above Unmatched since
        # these are older, already-flagged items to deal with first.
        self.saved_products_frame, saved_content = self._make_collapsible_panel(
            self.left_paned, self.SAVED_PRODUCTS_DESC, on_toggle=self._update_left_pane_sizes
        )
        self.saved_products_tree = self._create_treeview(
            saved_content, expand=True, height=min(max(len(self.saved_products), 1), 10)
        )
        saved_action_frame = ttk.Frame(saved_content)
        saved_action_frame.pack(fill="x", pady=(6, 0))
        ttk.Button(
            saved_action_frame, text="Ignore Selected", command=self._on_ignore_saved_product
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            saved_action_frame, text="Reassign to Existing Code...", command=self._on_reassign_saved_product
        ).pack(side="left")
        self._enable_row_copy(self.saved_products_tree)

        self.unmatched_frame, unmatched_content = self._make_collapsible_panel(
            self.left_paned, self.UNMATCHED_LAST_WEEK_DESC, on_toggle=self._update_left_pane_sizes
        )
        self.unmatched_tree = self._create_treeview(
            unmatched_content, expand=True, height=min(max(len(unmatched_rows), 1), 15)
        )

        action_frame = ttk.Frame(unmatched_content)
        action_frame.pack(fill="x", pady=(6, 0))
        ttk.Button(action_frame, text="Ignore Selected", command=self._on_ignore_unmatched).pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(
            action_frame, text="Save for Next Order", command=self._on_save_unmatched_for_later
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            action_frame, text="Reassign to Existing Code...", command=self._on_reassign_unmatched
        ).pack(side="left")
        self._enable_row_copy(self.unmatched_tree)

        # Pack-review panel — populated later by _apply_neto_stock_data() with products
        # sold in multi-unit packs, pulled out of the main table for manual checking.
        self.pack_review_frame, pack_review_content = self._make_collapsible_panel(
            self.left_paned, self.PACK_REVIEW_DESC, on_toggle=self._update_left_pane_sizes
        )
        self.pack_review_tree = self._create_treeview(pack_review_content, expand=True, height=8)
        pack_action_frame = ttk.Frame(pack_review_content)
        pack_action_frame.pack(fill="x", pady=(6, 0))
        ttk.Button(
            pack_action_frame, text="Ignore Selected", command=self._on_ignore_pack_review
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            pack_action_frame, text="Confirm & Return to Main Table",
            command=self._on_confirm_pack_review,
        ).pack(side="left")
        self._enable_row_copy(
            self.pack_review_tree,
            extra_commands=[("Change Recommended Packs...", self._on_change_recommended_packs)],
        )

        self.main_table_frame = ttk.Frame(self.paned)
        self.tree = self._create_treeview(self.main_table_frame)
        main_iids = self._populate_tree(
            self.tree,
            ["Product Code", "Product Name", "Qty Outstanding", self.NETO_QTY_SOLD_HEADER, self.PACKS_TO_ORDER_HEADER],
            matched_rows,
        )
        self.main_row_iid = {
            str(code): iid
            for (code, _name, _qty, _sold, _packs), iid in zip(matched_rows, main_iids)
        }
        self.main_code_pad_index = {
            str(code): i for i, (code, *_rest) in enumerate(matched_rows)
        }
        self._enable_row_copy(self.tree)

        self._set_saved_products_rows(self.saved_products)
        self._set_unmatched_rows(
            unmatched_rows, self.UNMATCHED_LAST_WEEK_DESC,
            qty_header="Qty Outstanding (Last Week)", target_column="Qty Outstanding",
        )
        self._set_pack_review_rows([])

        self.showing_outstanding_view = True

        status = f"Showing outstanding qty for {os.path.basename(self.this_week_pad_path.get())}"
        if unmatched_rows:
            status += f" — {len(unmatched_rows)} unmatched product code(s)"
        self.status_var.set(status)
        self._build_bottom_frame()

    def _set_unmatched_rows(self, rows, description, qty_header="Qty Outstanding", target_column="Qty Outstanding"):
        """(Re)populate the Unmatched Product Codes panel — used both for last week's
        unmatched codes (before Compare Orders) and for Neto-sold codes not found in this
        week's table (after Fetch Stock from Neto). target_column records which main-table
        column each row's qty represents, so Reassign writes to the right place."""
        self.unmatched_frame.gw_set_description(description)
        iids = self._populate_tree(
            self.unmatched_tree, ["Product Code", "Product Name", qty_header], rows
        )
        self.unmatched_row_data = {
            iid: {"code": code, "name": name, "qty": qty}
            for iid, (code, name, qty) in zip(iids, rows)
        }
        self.unmatched_target_column = target_column
        self._sync_paned_layout()

    def _set_pack_review_rows(self, rows):
        """(Re)populate the Pack Review panel with products sold in packs of more than
        one where more than 1 unit sold on Neto — see _apply_neto_stock_data().
        Row tuples: (code, name, qty_outstanding, qty_sold, sellable_stock_display,
        units_per_pack, recommended_packs)."""
        headers = [
            "Product Code", "Product Name", "Qty Outstanding", self.NETO_QTY_SOLD_HEADER,
            self.SELLABLE_STOCK_HEADER, self.QTY_IN_PACK_HEADER, self.RECOMMENDED_PACKS_HEADER,
        ]
        iids = self._populate_tree(self.pack_review_tree, headers, rows)
        self.pack_review_row_data = {
            iid: {"code": code, "name": name}
            for iid, (code, name, *_rest) in zip(iids, rows)
        }
        self._sync_paned_layout()

    def _current_pack_review_rows(self):
        """Read the Pack Review panel's current contents back out into row-tuple form
        (matching what _set_pack_review_rows expects) — used when a Reassign pushes a
        code over the pack-review threshold and it needs to be appended to the panel
        without dropping what's already there."""
        if not self.pack_review_tree:
            return []
        rows = []
        for iid in getattr(self.pack_review_tree, "gw_all_iids", ()):
            if not self.pack_review_tree.exists(iid):
                continue
            rows.append((
                self.pack_review_tree.set(iid, "Product Code"),
                self.pack_review_tree.set(iid, "Product Name"),
                self.pack_review_tree.set(iid, "Qty Outstanding"),
                self.pack_review_tree.set(iid, self.NETO_QTY_SOLD_HEADER),
                self.pack_review_tree.set(iid, self.SELLABLE_STOCK_HEADER),
                self.pack_review_tree.set(iid, self.QTY_IN_PACK_HEADER),
                self.pack_review_tree.set(iid, self.RECOMMENDED_PACKS_HEADER),
            ))
        return rows

    def _on_ignore_pack_review(self):
        """Drop the selected Pack Review lines without ordering anything — for when
        current stock is fine and no packs are needed this week."""
        if not self.pack_review_tree:
            return
        selected = self.pack_review_tree.selection()
        if not selected:
            messagebox.showinfo("No Selection", "Select one or more rows to ignore.")
            return

        for iid in selected:
            self.pack_review_tree.delete(iid)
            self.pack_review_row_data.pop(iid, None)

        count = len(selected)
        self.status_var.set(
            f"Ignored {count} pack-based line{'s' if count != 1 else ''} — no packs will be ordered."
        )

    def _on_confirm_pack_review(self):
        """Confirm the selected Pack Review lines as checked and return them to the
        main table, carrying the (possibly right-click-adjusted) Recommended Packs
        value into the main table's Packs to Order column."""
        if not self.pack_review_tree or not self.tree:
            return
        selected = self.pack_review_tree.selection()
        if not selected:
            messagebox.showinfo("No Selection", "Select one or more rows to confirm.")
            return

        for iid in selected:
            self._insert_main_row(
                self.pack_review_tree.set(iid, "Product Code"),
                self.pack_review_tree.set(iid, "Product Name"),
                self.pack_review_tree.set(iid, "Qty Outstanding"),
                self.pack_review_tree.set(iid, self.NETO_QTY_SOLD_HEADER),
                self.pack_review_tree.set(iid, self.RECOMMENDED_PACKS_HEADER),
            )
            self.pack_review_tree.delete(iid)
            self.pack_review_row_data.pop(iid, None)

        # Reattach main-table rows in gw_all_iids order so the returned lines land in
        # their original pad position rather than being tacked onto the end.
        self._filter_tree(self.tree)

        count = len(selected)
        self.status_var.set(
            f"Confirmed {count} pack-based line{'s' if count != 1 else ''} back to the main table."
        )

    def _insert_main_row(self, code, name, qty_outstanding, qty_sold, packs_to_order):
        """Put a row (removed earlier by _apply_neto_stock_data or a Reassign) back
        into the main table at its original order-pad position. The tree itself is
        appended to; ordering is restored by placing the new iid at the right spot in
        gw_all_iids and letting _filter_tree reattach everything in that order."""
        key = str(code)
        iid = self.tree.insert(
            "", "end",
            values=[str(code), str(name), str(qty_outstanding), str(qty_sold), str(packs_to_order)],
        )
        self.main_row_iid[key] = iid

        all_iids = list(getattr(self.tree, "gw_all_iids", []))
        pad_index = self.main_code_pad_index.get(key)
        insert_at = len(all_iids)
        if pad_index is not None:
            for i, existing in enumerate(all_iids):
                if not self.tree.exists(existing):
                    continue
                existing_idx = self.main_code_pad_index.get(
                    str(self.tree.set(existing, "Product Code"))
                )
                if existing_idx is not None and existing_idx > pad_index:
                    insert_at = i
                    break
        all_iids.insert(insert_at, iid)
        self.tree.gw_all_iids = all_iids

    def _on_change_recommended_packs(self):
        """Right-click override for the stock-based recommendation — e.g. order 2
        packs because demand is unusually high, or 0 because stock is on its way.
        Applies one number to every ticked line, so a batch of low-stock paints can
        be set in one go."""
        if not self.pack_review_tree:
            return
        selected = self.pack_review_tree.selection()
        if not selected:
            messagebox.showinfo("No Selection", "Tick one or more lines to change their recommended packs.")
            return

        first = selected[0]
        try:
            current = int(self.pack_review_tree.set(first, self.RECOMMENDED_PACKS_HEADER) or 0)
        except ValueError:
            current = 0

        if len(selected) == 1:
            code = self.pack_review_tree.set(first, "Product Code")
            name = self.pack_review_tree.set(first, "Product Name")
            prompt = f"Packs to order for {code}{' — ' + name if name else ''}:"
        else:
            prompt = f"Packs to order for the {len(selected)} ticked lines:"

        new_value = simpledialog.askinteger(
            "Change Recommended Packs", prompt,
            parent=self.root, initialvalue=current, minvalue=0,
        )
        if new_value is None:
            return

        for iid in selected:
            self.pack_review_tree.set(iid, self.RECOMMENDED_PACKS_HEADER, new_value)

        if len(selected) == 1:
            self.status_var.set(f"Recommended packs for {code} set to {new_value}.")
        else:
            self.status_var.set(f"Recommended packs set to {new_value} for {len(selected)} lines.")

    def _set_saved_products_rows(self, rows):
        """(Re)populate the Carried Over From Last Run panel from self.saved_products —
        entries saved via 'Save for Next Order' on the Unmatched panel, persisted in
        saved_products.json (see _load_saved_products/_write_saved_products)."""
        headers = ["Product Code", "Product Name", "Qty", "Saved On"]
        data_rows = [
            (item.get("code", ""), item.get("name", ""), item.get("qty", 0), item.get("saved_date", ""))
            for item in rows
        ]
        iids = self._populate_tree(self.saved_products_tree, headers, data_rows)
        self.saved_products_row_data = dict(zip(iids, rows))
        self._sync_paned_layout()

    def _sync_paned_layout(self):
        """Rebuild which left-column panels are attached, in a fixed order: Carried Over
        From Last Run, Unmatched Product Codes, then Pack Review. Unmatched is always
        shown; Carried Over and Pack Review only become relevant once Fetch Stock from
        Neto has run at least once (self.neto_fetch_done) — before that, the left column
        stays focused on just Unmatched Product Codes. Each attached panel is
        individually collapsible (see _make_collapsible_panel) rather than being
        automatically hidden based on whether it currently has data. Rebuilding from
        scratch each time is simpler than trying to insert a pane at a specific index."""
        if not self.paned or not self.left_paned:
            return

        for frame in (self.saved_products_frame, self.unmatched_frame, self.pack_review_frame):
            if frame and str(frame) in self.left_paned.panes():
                self.left_paned.forget(frame)

        if self.neto_fetch_done and self.saved_products_frame:
            self.left_paned.add(self.saved_products_frame, weight=1)
        if self.unmatched_frame:
            self.left_paned.add(self.unmatched_frame, weight=1)
        if self.neto_fetch_done and self.pack_review_frame:
            self.left_paned.add(self.pack_review_frame, weight=1)

        if self.left_paned and str(self.left_paned) not in self.paned.panes():
            # weight=0 keeps the left column at a fixed, narrow width — it doesn't
            # grow when the window is resized. All extra space goes to the main
            # table (weight=1) so it stays the dominant, most obvious part of the view.
            self.paned.add(self.left_paned, weight=0)
        if self.main_table_frame and str(self.main_table_frame) not in self.paned.panes():
            self.paned.add(self.main_table_frame, weight=1)

        if self.main_table_frame:
            # Force the initial split narrow on the left instead of the ~50/50 default
            # ttk.PanedWindow would otherwise give each pane's natural requested width.
            self.paned.update_idletasks()
            self.paned.sashpos(0, self.LEFT_PANE_WIDTH)

        self._update_left_pane_sizes()

    def _update_left_pane_sizes(self):
        """Recompute the left column's inner sash positions so a collapsed panel shrinks
        down to just its header height instead of keeping its old pixel allocation and
        leaving that space blank. Remaining height is split evenly across whichever
        panels are still expanded."""
        if not self.left_paned:
            return

        attached = [
            frame for frame in (self.saved_products_frame, self.unmatched_frame, self.pack_review_frame)
            if frame and str(frame) in self.left_paned.panes()
        ]
        if len(attached) < 2:
            return

        self.left_paned.update_idletasks()
        total_height = self.left_paned.winfo_height()
        if total_height <= 1:
            return

        heights = []
        expanded_idx = []
        for i, frame in enumerate(attached):
            if getattr(frame, "gw_is_collapsed", lambda: False)():
                heights.append(max(frame.winfo_reqheight(), 1))
            else:
                heights.append(None)
                expanded_idx.append(i)

        collapsed_total = sum(h for h in heights if h is not None)
        remaining = max(total_height - collapsed_total, 0)
        share = remaining // len(expanded_idx) if expanded_idx else 0
        for i in expanded_idx:
            heights[i] = share

        cumulative = 0
        for i in range(len(attached) - 1):
            cumulative += heights[i]
            try:
                self.left_paned.sashpos(i, cumulative)
            except tk.TclError:
                pass

    def _copy_selected_codes(self, tree):
        selected = tree.selection()
        if not selected:
            return

        codes = [tree.set(iid, "Product Code") for iid in selected]
        self.root.clipboard_clear()
        self.root.clipboard_append("\n".join(codes))

        if len(codes) == 1:
            self.status_var.set(f"Copied product code {codes[0]} to clipboard.")
        else:
            self.status_var.set(f"Copied {len(codes)} product codes to clipboard.")

    def _select_all_rows(self, tree):
        """Select every currently visible row — respects an active search filter, since
        tree.get_children() only returns attached (i.e. visible) top-level items."""
        tree.selection_set(tree.get_children())
        return "break"

    def _has_check_column(self, tree):
        columns = tree["columns"]
        return bool(columns) and columns[0] == self.CHECK_HEADER

    def _on_tree_click(self, tree, event):
        """Clicks on the checkbox column: a row's box toggles just that row in/out of
        the selection (no Ctrl needed); the column heading ticks all visible rows, or
        unticks them if everything is already ticked. Clicks anywhere else fall
        through to ttk's normal selection behaviour."""
        if not self._has_check_column(tree) or tree.identify_column(event.x) != "#1":
            return
        region = tree.identify_region(event.x, event.y)

        if region == "heading":
            visible = tree.get_children()
            if visible and set(visible) <= set(tree.selection()):
                tree.selection_remove(*visible)
            elif visible:
                tree.selection_set(visible)
            return "break"

        if region == "cell":
            iid = tree.identify_row(event.y)
            if not iid:
                return
            if iid in tree.selection():
                tree.selection_remove(iid)
            else:
                tree.selection_add(iid)
            return "break"

    def _sync_check_marks(self, tree):
        """Redraw ☐/☑ to mirror the tree's current selection — runs on every
        <<TreeviewSelect>>, so ticks stay correct however the selection was made
        (box click, Ctrl/Shift-click, Ctrl+A, or an action removing rows)."""
        if not self._has_check_column(tree):
            return
        selected = set(tree.selection())
        for iid in getattr(tree, "gw_all_iids", ()):
            if tree.exists(iid):
                tree.set(iid, self.CHECK_HEADER, self.CHECKED if iid in selected else self.UNCHECKED)

    def _enable_row_copy(self, tree, extra_commands=None):
        """Right-click context menu with Select All / Copy on every table;
        extra_commands ([(label, callback), ...]) appends table-specific actions —
        e.g. Change Recommended Packs... on the Pack Review panel."""
        menu = tk.Menu(tree, tearoff=0)
        menu.add_command(label="Select All", command=lambda: self._select_all_rows(tree))
        menu.add_command(label="Copy Product Code", command=lambda: self._copy_selected_codes(tree))
        if extra_commands:
            menu.add_separator()
            for label, callback in extra_commands:
                menu.add_command(label=label, command=callback)

        def show_context_menu(event):
            iid = tree.identify_row(event.y)
            if iid and iid not in tree.selection():
                tree.selection_set(iid)
            menu.tk_popup(event.x_root, event.y_root)

        tree.bind("<Control-c>", lambda e: self._copy_selected_codes(tree))
        tree.bind("<Command-c>", lambda e: self._copy_selected_codes(tree))
        tree.bind("<Button-3>", show_context_menu)
        tree.bind("<Button-2>", show_context_menu)
        tree.bind("<Control-Button-1>", show_context_menu)

    def _on_ignore_unmatched(self):
        if not self.unmatched_tree:
            return
        selected = self.unmatched_tree.selection()
        if not selected:
            messagebox.showinfo("No Selection", "Select one or more rows to ignore.")
            return

        for iid in selected:
            self.unmatched_tree.delete(iid)
            self.unmatched_row_data.pop(iid, None)

        self._sync_paned_layout()

    def _on_save_unmatched_for_later(self):
        """Save the selected unmatched product(s) to saved_products.json for follow-up
        on a future run, and remove them from the Unmatched panel immediately — saving
        counts as resolving them for this session, same as Ignore or Reassign."""
        if not self.unmatched_tree:
            return
        selected = self.unmatched_tree.selection()
        if not selected:
            messagebox.showinfo("No Selection", "Select one or more rows to save for next time.")
            return

        today_str = date.today().isoformat()
        saved_lookup = {str(item.get("code")): item for item in self.saved_products}

        for iid in selected:
            row = self.unmatched_row_data.get(iid)
            if not row:
                continue
            code = str(row["code"])
            try:
                qty = int(row["qty"] or 0)
            except (TypeError, ValueError):
                qty = 0

            if code in saved_lookup:
                existing = saved_lookup[code]
                try:
                    existing_qty = int(existing.get("qty") or 0)
                except (TypeError, ValueError):
                    existing_qty = 0
                existing["qty"] = existing_qty + qty
                existing["name"] = row.get("name") or existing.get("name", "")
                existing["saved_date"] = today_str
            else:
                saved_lookup[code] = {
                    "code": code, "name": row.get("name", ""), "qty": qty, "saved_date": today_str,
                }

            self.unmatched_tree.delete(iid)
            self.unmatched_row_data.pop(iid, None)

        self.saved_products = list(saved_lookup.values())
        self._write_saved_products()
        self._set_saved_products_rows(self.saved_products)
        self._sync_paned_layout()

        count = len(selected)
        self.status_var.set(
            f"Saved {count} product{'s' if count != 1 else ''} to Carried Over From Last Run."
        )

    def _on_ignore_saved_product(self):
        """Ignore selected products from the Carried Over From Last Run panel once
        they've been dealt with, and drop them from saved_products.json for good —
        same idea as Ignore Selected on the Unmatched Product Codes panel."""
        if not self.saved_products_tree:
            return
        selected = self.saved_products_tree.selection()
        if not selected:
            messagebox.showinfo("No Selection", "Select one or more rows to ignore.")
            return

        codes_to_remove = {
            str(self.saved_products_row_data[iid].get("code"))
            for iid in selected if iid in self.saved_products_row_data
        }
        self.saved_products = [
            item for item in self.saved_products if str(item.get("code")) not in codes_to_remove
        ]
        self._write_saved_products()
        self._set_saved_products_rows(self.saved_products)

        count = len(selected)
        self.status_var.set(f"Ignored {count} saved product{'s' if count != 1 else ''}.")

    def _on_reassign_saved_product(self):
        """Reassign a Carried Over From Last Run product onto an existing product code
        in this week's main table — same dialog as reassigning an unmatched code."""
        if not self.saved_products_tree:
            return
        selected = self.saved_products_tree.selection()
        if len(selected) != 1:
            messagebox.showinfo("Select One Row", "Select exactly one saved product to reassign.")
            return

        iid = selected[0]
        row = self.saved_products_row_data.get(iid)
        if not row:
            return

        code = str(row.get("code"))
        name = row.get("name", "")
        try:
            qty = int(row.get("qty") or 0)
        except (TypeError, ValueError):
            qty = 0

        def on_resolved():
            self.saved_products = [
                item for item in self.saved_products if str(item.get("code")) != code
            ]
            self._write_saved_products()
            self._set_saved_products_rows(self.saved_products)

        # Carried-over items predate this session, so there's no "Qty Sold on Neto"
        # context to preserve — always reassign onto Qty Outstanding.
        self._open_reassign_dialog(code, name, qty, "Qty Outstanding", on_resolved)

    def _on_reassign_unmatched(self):
        if not self.unmatched_tree:
            return
        selected = self.unmatched_tree.selection()
        if len(selected) != 1:
            messagebox.showinfo("Select One Row", "Select exactly one unmatched product code to reassign.")
            return

        iid = selected[0]
        row = self.unmatched_row_data.get(iid)
        if not row:
            return

        def on_resolved():
            self.unmatched_tree.delete(iid)
            self.unmatched_row_data.pop(iid, None)
            self._sync_paned_layout()

        self._open_reassign_dialog(
            row["code"], row["name"], row["qty"], self.unmatched_target_column, on_resolved
        )

    def _bind_autocomplete(self, combo, all_values):
        def on_keyrelease(event):
            if event.keysym in ("Up", "Down", "Return", "Escape", "Tab"):
                return

            typed = combo.get()
            if typed == "":
                filtered = all_values
            else:
                typed_lower = typed.lower()
                filtered = [v for v in all_values if typed_lower in v.lower()]

            combo["values"] = filtered

            if typed and filtered:
                combo.event_generate("<Down>")
                combo.focus_set()
                combo.icursor(tk.END)

        combo.bind("<KeyRelease>", on_keyrelease)

    def _open_reassign_dialog(
        self, old_code, old_name, old_qty, target_column="Qty Outstanding", on_resolved=None
    ):
        codes = sorted(self.main_row_iid.keys())
        if not codes:
            messagebox.showinfo("No Product Codes", "There are no product codes in this week's table to reassign to.")
            return

        is_neto_mode = target_column == self.NETO_QTY_SOLD_HEADER
        qty_label_text = f"{self.NETO_QTY_SOLD_HEADER}:" if is_neto_mode else "Qty Outstanding:"

        dialog = tk.Toplevel(self.root)
        dialog.title("Reassign Product Code")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        unmatched_label = f"Reassigning: {old_code}"
        if old_name:
            unmatched_label += f" — {old_name}"
        ttk.Label(dialog, text=unmatched_label).grid(
            row=0, column=0, columnspan=2, sticky="w", padx=10, pady=(10, 6)
        )

        ttk.Label(dialog, text="New Product Code:").grid(row=1, column=0, sticky="w", padx=10, pady=4)
        code_var = tk.StringVar()
        combo = ttk.Combobox(dialog, textvariable=code_var, values=codes, width=28)
        combo.grid(row=1, column=1, padx=10, pady=4)
        self._bind_autocomplete(combo, codes)

        ttk.Label(dialog, text=qty_label_text).grid(row=2, column=0, sticky="w", padx=10, pady=4)
        qty_var = tk.StringVar(value=str(old_qty))
        qty_entry = ttk.Entry(dialog, textvariable=qty_var, width=15)
        qty_entry.grid(row=2, column=1, sticky="w", padx=10, pady=4)

        def on_code_selected(event=None):
            # Suggest the combined total (target row's current qty + this unmatched row's
            # qty) rather than silently overwriting whatever the target already had.
            main_iid = self.main_row_iid.get(code_var.get().strip())
            if main_iid is None:
                return
            try:
                current_qty = int(self.tree.set(main_iid, target_column) or 0)
            except ValueError:
                current_qty = 0
            qty_var.set(str(current_qty + old_qty))

        combo.bind("<<ComboboxSelected>>", on_code_selected)

        btn_frame = ttk.Frame(dialog)
        btn_frame.grid(row=3, column=0, columnspan=2, pady=(10, 10))

        def on_confirm():
            new_code = code_var.get().strip()
            if not new_code:
                messagebox.showwarning("Missing Code", "Please select a product code.", parent=dialog)
                return
            try:
                new_qty = int(qty_var.get().strip())
            except ValueError:
                messagebox.showwarning("Invalid Qty", "Qty must be a whole number.", parent=dialog)
                return

            main_iid = self.main_row_iid.get(new_code)
            if main_iid is None:
                messagebox.showwarning(
                    "Not Found", "Selected product code was not found in the table.", parent=dialog
                )
                return

            if is_neto_mode:
                # Reassigning a Neto-sold code onto an existing row — write to Qty Sold
                # on Neto (not Qty Outstanding) and recompute Packs to Order. If that
                # pushes it over the pack-review threshold, move it to the Pack Review
                # panel just like a normal Fetch Stock from Neto would.
                qty_in_pack = self.qty_in_pack_lookup.get(new_code, 1)
                if not isinstance(qty_in_pack, int) or qty_in_pack <= 0:
                    qty_in_pack = 1
                packs_to_order = math.ceil(new_qty / qty_in_pack)

                if qty_in_pack > 1 and new_qty > 1:
                    qty_outstanding = self.tree.set(main_iid, "Qty Outstanding")
                    product_name = self.tree.set(main_iid, "Product Name")
                    self.tree.delete(main_iid)
                    if hasattr(self.tree, "gw_all_iids"):
                        self.tree.gw_all_iids = [i for i in self.tree.gw_all_iids if i != main_iid]
                    del self.main_row_iid[new_code]

                    # Stock was scraped under the code the product sold as on Neto (the
                    # old code) — fall back to the new code, then to unknown (blank
                    # stock, conservative 1-pack recommendation).
                    stock = self.neto_stock_lookup.get(str(old_code)) or self.neto_stock_lookup.get(new_code)
                    if stock is not None:
                        stock_display = self._format_stock_display(*stock)
                        recommended = self._recommended_packs(stock[0])
                    else:
                        stock_display = ""
                        recommended = 1

                    pack_rows = self._current_pack_review_rows()
                    pack_rows.append((
                        new_code, product_name, qty_outstanding, new_qty,
                        stock_display, qty_in_pack, recommended,
                    ))
                    self._set_pack_review_rows(pack_rows)
                else:
                    self.tree.set(main_iid, self.NETO_QTY_SOLD_HEADER, new_qty)
                    self.tree.set(main_iid, self.PACKS_TO_ORDER_HEADER, packs_to_order)
            else:
                self.tree.set(main_iid, "Qty Outstanding", new_qty)
                self.outstanding_lookup[new_code] = new_qty

            if on_resolved:
                on_resolved()

            dialog.destroy()

        ttk.Button(btn_frame, text="Confirm", command=on_confirm).pack(side="left", padx=6)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side="left", padx=6)

    def _on_fetch_neto(self):
        if self.unmatched_row_data:
            count = len(self.unmatched_row_data)
            messagebox.showwarning(
                "Unmatched Product Codes",
                f"There {'is' if count == 1 else 'are'} still {count} unmatched product code"
                f"{'' if count == 1 else 's'} that haven't been resolved.\n\n"
                "Please use \"Ignore Selected\", \"Save for Next Order\", or \"Reassign to Existing "
                "Code...\" to resolve them before fetching stock from Neto.",
            )
            return

        self.neto_btn.config(state="disabled")
        self._set_neto_checking(True)
        self.status_var.set(
            f"Checking product barcodes on Neto (orders placed since {self.neto_date_var.get()}) — please wait..."
        )

        thread = threading.Thread(target=self._run_neto_scraper, daemon=True)
        thread.start()

    def _set_neto_checking(self, active):
        """Make it obvious the app is actively checking product barcodes against Neto,
        instead of relying on easy-to-miss status bar text alone."""
        if not self.status_label or not self.neto_progress:
            return
        if active:
            self.status_label.config(style="NetoChecking.TLabel")
            self.neto_progress.pack(side="left", padx=(10, 0))
            self.neto_progress.start(10)
        else:
            self.neto_progress.stop()
            self.neto_progress.pack_forget()
            self.status_label.config(style="TLabel")

    def _set_neto_login_waiting(self, waiting, text):
        """Distinct visual state for 'stopped and waiting on you to log in' vs. the normal
        'actively checking Neto' state, so it's obvious action is needed on your end."""
        if not self.status_label:
            return
        self.status_label.config(style="NetoLoginWait.TLabel" if waiting else "NetoChecking.TLabel")
        self.status_var.set(text)

    def _run_neto_scraper(self):
        script_path = os.path.join(self.script_dir, "neto_scraper.py")
        from_date_str = self._format_date(self.neto_from_date)
        login_prompt_shown = False
        try:
            # "-u" forces the child to run with unbuffered stdout. Without it, Python
            # fully buffers stdout when it's a pipe (not a terminal), so print() lines
            # like "Not logged in to Neto yet." can sit in the child's buffer for a
            # while before we ever see them — the Chrome window updates live, but the
            # status bar here would lag behind and look stuck on stale text.
            process = subprocess.Popen(
                [sys.executable, "-u", script_path, "--from-date", from_date_str],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
            )

            self.root.after(0, lambda: self.status_var.set("Checking product barcodes on Neto..."))

            output = []
            for line in process.stdout:
                stripped = line.strip()
                output.append(stripped)

                if stripped.startswith(("Not logged in to Neto", "Please log in using the Chrome window")):
                    display = "Waiting for you to log in to Neto — a Chrome window is open for you to sign in."
                    self.root.after(0, lambda d=display: self._set_neto_login_waiting(True, d))
                elif stripped.startswith("Still waiting for login"):
                    suffix = stripped.split("...", 1)[1].strip() if "..." in stripped else ""
                    display = f"Still waiting for you to log in to Neto... {suffix}".strip()
                    self.root.after(0, lambda d=display: self._set_neto_login_waiting(True, d))
                elif stripped.startswith("Login detected"):
                    display = "Login detected — resuming stock check on Neto..."
                    self.root.after(0, lambda d=display: self._set_neto_login_waiting(False, d))
                else:
                    display = f"Checking Neto: {stripped}" if stripped else "Checking product barcodes on Neto..."
                    self.root.after(0, lambda d=display: self.status_var.set(d))

                if not login_prompt_shown and stripped.startswith("Not logged in to Neto"):
                    login_prompt_shown = True
                    self.root.after(0, lambda: messagebox.showinfo(
                        "Login Required",
                        "You're not logged in to Neto yet.\n\n"
                        "A Chrome window has opened — please log in there. This will "
                        "continue automatically once it detects you're logged in.",
                    ))

            process.wait()

            if process.returncode == 0:
                self.root.after(0, lambda: self.status_var.set("Neto scraper finished. Applying stock data..."))
                self.root.after(0, self._load_and_apply_neto_stock_data)
            else:
                last_line = next((l for l in reversed(output) if l), "")
                self.root.after(0, lambda: self.status_var.set(f"Scraper exited with error (code {process.returncode})"))
                self.root.after(0, lambda msg=last_line, code=process.returncode: messagebox.showerror(
                    "Neto Scraper Failed",
                    msg or f"Scraper exited with error (code {code}).",
                ))

        except Exception as e:
            self.root.after(0, lambda: self.status_var.set(f"Error: {e}"))

        self.root.after(0, lambda: self.neto_btn.config(state="normal"))
        self.root.after(0, lambda: self._set_neto_checking(False))

    def _load_and_apply_neto_stock_data(self):
        """Runs on the main thread after the scraper exits successfully. Reads the demand
        data it wrote out (one entry per SKU sold on Neto) and folds it into the table."""
        json_path = os.path.join(self.script_dir, "sales_order_demand.json")
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                sku_summary = json.load(f)
        except Exception as e:
            messagebox.showerror(
                "Failed to Read Neto Data",
                f"Neto scraper finished, but sales_order_demand.json couldn't be read:\n{e}",
            )
            self.status_var.set("Neto scraper finished, but stock data couldn't be applied.")
            return

        self._apply_neto_stock_data(sku_summary)

    @classmethod
    def _recommended_packs(cls, stock_available):
        """Stock-based pack recommendation: pack products are ordered to keep shelf
        stock topped up, not to match units sold — sellable stock already accounts for
        what's reserved by pending orders. At or below LOW_STOCK_THRESHOLD sellable
        units, recommend 1 pack; above it, nothing needs ordering. The user can
        override per line via right-click (see _on_change_recommended_packs)."""
        return 1 if stock_available <= cls.LOW_STOCK_THRESHOLD else 0

    @staticmethod
    def _format_stock_display(stock_available, stock_on_hand):
        """Mirror Neto's own 'Stock: 4 (5)' style — sellable stock first, physical
        on-hand stock in brackets only when it differs (i.e. some is reserved)."""
        if stock_on_hand is not None and stock_on_hand != stock_available:
            return f"{stock_available} ({stock_on_hand})"
        return str(stock_available)

    def _apply_neto_stock_data(self, sku_summary):
        """For every SKU sold on Neto since last Tuesday, show the qty sold in its own
        column for review, plus how many whole packs that works out to (rounded up —
        e.g. 12 units sold at 6/pack = 2 packs). Qty Outstanding is left exactly as it
        was before the fetch — deliberately NOT auto-added here, so it can be checked
        first. Products sold in packs of more than one, with more than 1 unit sold on
        Neto, are pulled OUT of the main table and into the Pack Review panel instead;
        there the recommendation is based on current sellable stock rather than qty
        sold (see _recommended_packs), and each line is confirmed or ignored manually.
        SKUs that don't exist in this week's table at all are surfaced in the
        Unmatched Product Codes panel."""
        self.neto_fetch_done = True
        unmatched = []
        pack_review_rows = []
        updated = 0

        for entry in sku_summary:
            sku = str(entry.get("sku", "")).strip()
            if not sku:
                continue
            qty_needed = int(entry.get("total_qty_needed") or 0)
            if qty_needed <= 0:
                continue
            product_name = entry.get("product_name", "")

            try:
                stock_available = int(entry.get("stock_available_to_sell", entry.get("stock", 0)) or 0)
            except (TypeError, ValueError):
                stock_available = 0
            try:
                raw_on_hand = entry.get("stock_on_hand")
                stock_on_hand = int(raw_on_hand) if raw_on_hand is not None else None
            except (TypeError, ValueError):
                stock_on_hand = None
            self.neto_stock_lookup[sku] = (stock_available, stock_on_hand)

            main_iid = self.main_row_iid.get(sku)
            if main_iid is None:
                unmatched.append((sku, product_name, qty_needed))
                continue

            qty_in_pack = self.qty_in_pack_lookup.get(sku, 1)
            if not isinstance(qty_in_pack, int) or qty_in_pack <= 0:
                qty_in_pack = 1
            packs_to_order = math.ceil(qty_needed / qty_in_pack)

            if qty_in_pack > 1 and qty_needed > 1:
                # Needs manual review — pull it out of the main table entirely rather
                # than leaving a possibly-misleading auto-computed number in place.
                qty_outstanding = self.tree.set(main_iid, "Qty Outstanding")
                pack_review_rows.append((
                    sku, product_name, qty_outstanding, qty_needed,
                    self._format_stock_display(stock_available, stock_on_hand),
                    qty_in_pack, self._recommended_packs(stock_available),
                ))
                self.tree.delete(main_iid)
                if hasattr(self.tree, "gw_all_iids"):
                    self.tree.gw_all_iids = [i for i in self.tree.gw_all_iids if i != main_iid]
                del self.main_row_iid[sku]
            else:
                # Qty Outstanding is intentionally untouched here — see docstring.
                self.tree.set(main_iid, self.NETO_QTY_SOLD_HEADER, qty_needed)
                self.tree.set(main_iid, self.PACKS_TO_ORDER_HEADER, packs_to_order)

            updated += 1

        self._set_unmatched_rows(
            unmatched, self.UNMATCHED_NETO_DESC,
            qty_header="Qty Sold on Neto", target_column=self.NETO_QTY_SOLD_HEADER,
        )
        self._set_pack_review_rows(pack_review_rows)

        status = f"Fetched stock from Neto — {updated} product(s) matched"
        if pack_review_rows:
            status += f", {len(pack_review_rows)} moved to Pack Review"
        if unmatched:
            status += f", {len(unmatched)} sold on Neto but not found in this week's table"
        else:
            status += "."
        self.status_var.set(status)

    def _populate_tree(self, tree, headers, data_rows):
        # tree.get_children() only returns currently-attached (visible) items — anything
        # hidden by an active search filter is detached, not deleted, so it has to be
        # cleaned up explicitly here or it'd silently pile up on every repopulate.
        for iid in getattr(tree, "gw_all_iids", ()):
            if tree.exists(iid):
                tree.delete(iid)
        tree.delete(*tree.get_children())

        # Bulk-select checkbox column always comes first — callers pass headers/rows
        # without it and never need to know it exists (all cell access is by column
        # name, so prepending doesn't shift anything).
        headers = [self.CHECK_HEADER] + list(headers)

        tree["columns"] = headers
        for h in headers:
            tree.heading(h, text=h)
            min_w = self.COLUMN_MIN_WIDTHS.get(h, self.DEFAULT_COLUMN_MIN_WIDTH)
            tree.column(h, width=min_w, minwidth=min_w, stretch=True)

        iids = []
        for row in data_rows:
            values = [self.UNCHECKED] + [str(v) if v is not None else "" for v in row]
            iids.append(tree.insert("", "end", values=values))
        tree.gw_all_iids = iids

        # Recompute actual column widths any time the table is resized, so it always
        # fills the available width rather than leaving blank space on the right.
        tree.bind("<Configure>", lambda e, t=tree, hs=headers: self._resize_tree_columns(t, hs, e.width))
        tree.update_idletasks()
        self._resize_tree_columns(tree, headers, tree.winfo_width())

        # Re-apply whatever search text is already in the box (e.g. after resolving an
        # unmatched row triggers a repopulate) instead of silently clearing the filter.
        self._filter_tree(tree)

        return iids

    def _filter_tree(self, tree):
        """Show only rows whose Product Code or Product Name match the search box,
        using detach/reattach rather than delete/reinsert so item ids stay stable —
        main_row_iid and unmatched_row_data reference these ids directly."""
        search_var = getattr(tree, "gw_search_var", None)
        all_iids = getattr(tree, "gw_all_iids", None)
        if search_var is None or not all_iids:
            return

        query = search_var.get().strip().lower()
        columns = tree["columns"]
        has_code = "Product Code" in columns
        has_name = "Product Name" in columns

        tree.detach(*[iid for iid in all_iids if tree.exists(iid)])
        for iid in all_iids:
            if not tree.exists(iid):
                continue
            if not query:
                tree.reattach(iid, "", "end")
                continue
            code = str(tree.set(iid, "Product Code")).lower() if has_code else ""
            name = str(tree.set(iid, "Product Name")).lower() if has_name else ""
            if query in code or query in name:
                tree.reattach(iid, "", "end")

    def _resize_tree_columns(self, tree, headers, total_width):
        if not headers or total_width <= 1:
            return

        weights = [self.COLUMN_WEIGHTS.get(h, self.DEFAULT_COLUMN_WEIGHT) for h in headers]
        min_widths = [self.COLUMN_MIN_WIDTHS.get(h, self.DEFAULT_COLUMN_MIN_WIDTH) for h in headers]
        total_weight = sum(weights) or 1
        extra = max(total_width - sum(min_widths), 0)

        for h, weight, min_w in zip(headers, weights, min_widths):
            bonus = int(extra * weight / total_weight)
            try:
                tree.column(h, width=min_w + bonus)
            except tk.TclError:
                # Column may no longer exist if the tree got repopulated mid-resize.
                pass


def main():
    root = tk.Tk()
    GWOrderTool(root)
    root.mainloop()


if __name__ == "__main__":
    main()
